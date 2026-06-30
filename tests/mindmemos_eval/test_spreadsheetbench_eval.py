"""Tests for SpreadsheetBench evaluation helpers."""

from __future__ import annotations

import asyncio
import json
import tarfile
from pathlib import Path
from typing import Any

import pytest
from mindmemos_eval.skills.envs.spreadsheetbench import (
    EvolveOutcome,
    ReactAgentTools,
    SpreadsheetBenchCaseResult,
    SpreadsheetBenchEnv,
    compare_cell_value,
    compare_workbooks,
    generate_cell_names,
    prepare_data_root,
)

pytest.importorskip("openpyxl")

SKILL_RL_SPREADSHEET_SYSTEM_PROMPT = (
    "You are an expert spreadsheet assistant. Your working directory contains a "
    "source Excel file named 'input.xlsx'. Do NOT modify 'input.xlsx'. Instead, "
    "produce a new file 'output.xlsx' in the same directory that fully satisfies "
    "the user's request (start from a copy of 'input.xlsx' and apply your changes).\n"
    "Work by writing and running Python (openpyxl is available) through the shell "
    "tool — do not answer from memory. Inspect the sheets first, apply the changes, "
    "save to 'output.xlsx', and verify.\n"
    "IMPORTANT: 'output.xlsx' is graded by reading cached cell VALUES, with no "
    "formula recalculation. Write the final computed values into the target cells "
    "(not bare formulas), since an unevaluated formula reads back as empty.\n"
    "When you are done, stop without calling any tool."
)


def _make_workbook(path: Path, value: Any) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = value
    wb.save(path)
    wb.close()


def _make_dataset(root: Path) -> Path:
    data_root = root / "SpreadsheetBench"
    sheet_dir = data_root / "spreadsheetbench_verified_400" / "spreadsheet" / "1"
    split_dir = data_root / "spreadsheetbench_id_split" / "test"
    sheet_dir.mkdir(parents=True)
    split_dir.mkdir(parents=True)

    _make_workbook(sheet_dir / "1_init.xlsx", 1)
    _make_workbook(sheet_dir / "1_golden.xlsx", 42)
    (data_root / "spreadsheetbench_verified_400" / "dataset.json").write_text(
        json.dumps(
            [
                {
                    "id": "1",
                    "instruction": "Set A1 to 42.",
                    "spreadsheet_path": "spreadsheet/1",
                    "answer_position": "A1",
                    "answer_sheet": "Sheet1",
                    "instruction_type": "unit",
                }
            ]
        ),
        encoding="utf-8",
    )
    (split_dir / "items.json").write_text(json.dumps([{"id": "1"}]), encoding="utf-8")
    return data_root


def _make_all_dataset(root: Path, size: int = 3) -> Path:
    data_root = root / "SpreadsheetBench"
    verified = data_root / "spreadsheetbench_verified_400"
    records = []
    for index in range(1, size + 1):
        sheet_dir = verified / "spreadsheet" / str(index)
        sheet_dir.mkdir(parents=True)
        _make_workbook(sheet_dir / f"{index}_init.xlsx", index)
        _make_workbook(sheet_dir / f"{index}_golden.xlsx", index)
        records.append(
            {
                "id": str(index),
                "instruction": f"Leave A1 as {index}.",
                "spreadsheet_path": f"spreadsheet/{index}",
                "answer_position": "A1",
                "answer_sheet": "Sheet1",
                "instruction_type": "unit",
            }
        )
    (verified / "dataset.json").write_text(json.dumps(records), encoding="utf-8")
    return data_root


def _make_mixed_id_dataset(root: Path) -> Path:
    data_root = root / "SpreadsheetBench"
    verified = data_root / "spreadsheetbench_verified_400"
    records = []
    for record_id in ["10", "task_a", "2"]:
        sheet_dir = verified / "spreadsheet" / record_id
        sheet_dir.mkdir(parents=True)
        _make_workbook(sheet_dir / f"{record_id}_init.xlsx", 1)
        _make_workbook(sheet_dir / f"{record_id}_golden.xlsx", 1)
        records.append(
            {
                "id": record_id,
                "instruction": "Leave A1 unchanged.",
                "spreadsheet_path": f"spreadsheet/{record_id}",
                "answer_position": "A1",
                "answer_sheet": "Sheet1",
                "instruction_type": "unit",
            }
        )
    (verified / "dataset.json").write_text(json.dumps(records), encoding="utf-8")
    return data_root


def test_compare_workbooks_uses_spreadsheetbench_value_semantics(tmp_path: Path) -> None:
    gt = tmp_path / "gt.xlsx"
    pred = tmp_path / "pred.xlsx"
    _make_workbook(gt, 1.234)
    _make_workbook(pred, "1.23")

    ok, message = compare_workbooks(gt, pred, "Sheet1!A1")

    assert ok
    assert message == ""
    assert compare_cell_value("", None)
    assert generate_cell_names("A1:B2") == ["A1", "A2", "B1", "B2"]


def test_agent_tools_keep_access_inside_workdir(tmp_path: Path) -> None:
    tools = ReactAgentTools(tmp_path)

    assert "Successfully wrote" in tools.write("note.txt", "hello")
    assert tools.read("note.txt") == "hello"
    assert "outside the working directory" in tools.read("../secret.txt")
    assert "installing packages is disabled" in tools.shell(["pip install nope"])


@pytest.mark.asyncio
async def test_spreadsheetbench_env_runs_one_case_and_scores_success(tmp_path: Path) -> None:
    data_root = _make_dataset(tmp_path)
    run_dir = tmp_path / "run"

    async def fake_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        del tools
        if not any(message.get("role") == "tool" for message in messages):
            return {
                "role": "assistant",
                "content": None,
                "tool_calls": [
                    {
                        "id": "call_1",
                        "type": "function",
                        "function": {
                            "name": "shell",
                            "arguments": json.dumps(
                                {
                                    "commands": [
                                        "python - <<'PY'\n"
                                        "import openpyxl, shutil\n"
                                        "shutil.copyfile('input.xlsx', 'output.xlsx')\n"
                                        "wb = openpyxl.load_workbook('output.xlsx')\n"
                                        "ws = wb['Sheet1']\n"
                                        "ws['A1'] = 42\n"
                                        "wb.save('output.xlsx')\n"
                                        "wb.close()\n"
                                        "PY"
                                    ]
                                }
                            ),
                        },
                    }
                ],
            }
        return {"role": "assistant", "content": "Done."}

    trajectory_path = tmp_path / "custom" / "traces.jsonl"
    env = SpreadsheetBenchEnv(
        data_root=data_root,
        run_dir=run_dir,
        llm=fake_llm,
        max_turns=3,
        trajectory_path=trajectory_path,
    )
    result = await env.evaluate(split="test", show_progress=True)

    assert result.total == 1
    assert result.correct == 1
    assert result.accuracy == 1.0
    assert (run_dir / "test_summary.json").exists()
    case_result = result.results[0]
    assert case_result.finished
    assert case_result.score == 1.0
    assert Path(case_result.workdir, "output.xlsx").exists()
    assert trajectory_path.exists()
    assert json.loads(trajectory_path.read_text(encoding="utf-8").splitlines()[0])["messages"]


@pytest.mark.asyncio
async def test_spreadsheetbench_env_runs_baseline_concurrently(tmp_path: Path) -> None:
    data_root = _make_all_dataset(tmp_path, size=3)
    run_dir = tmp_path / "run"

    async def fake_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        del tools
        if not any(message.get("role") == "tool" for message in messages):
            return {
                "role": "assistant",
                "content": None,
                "tool_calls": [
                    {
                        "id": "call_1",
                        "type": "function",
                        "function": {
                            "name": "shell",
                            "arguments": json.dumps(
                                {
                                    "commands": [
                                        "python - <<'PY'\n"
                                        "import shutil\n"
                                        "shutil.copyfile('input.xlsx', 'output.xlsx')\n"
                                        "PY"
                                    ]
                                }
                            ),
                        },
                    }
                ],
            }
        return {"role": "assistant", "content": "Done."}

    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=run_dir, llm=fake_llm, max_turns=3)
    result = await env.evaluate(split="all", concurrency=2, show_progress=True)

    assert result.total == 3
    assert result.correct == 3
    assert result.accuracy == 1.0
    assert (run_dir / "all_trajectories.jsonl").exists()


def test_spreadsheetbench_env_loads_all_cases_without_split(tmp_path: Path) -> None:
    data_root = _make_all_dataset(tmp_path, size=3)

    async def never_called_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        raise AssertionError("LLM should not be called while loading cases")

    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=tmp_path / "run", llm=never_called_llm)
    cases = env.load_cases("all")
    shuffled_a = env.load_cases("all")
    shuffled_b = env.load_cases("all")

    import random

    random.Random(1447).shuffle(shuffled_a)
    random.Random(1447).shuffle(shuffled_b)

    assert [case.id for case in cases] == ["1", "2", "3"]
    assert [case.id for case in shuffled_a] == [case.id for case in shuffled_b]
    assert [case.id for case in shuffled_a] != ["1", "2", "3"]


def test_spreadsheetbench_env_sorts_mixed_ids(tmp_path: Path) -> None:
    data_root = _make_mixed_id_dataset(tmp_path)

    async def never_called_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        raise AssertionError("LLM should not be called while loading cases")

    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=tmp_path / "run", llm=never_called_llm)

    assert [case.id for case in env.load_cases("all")] == ["2", "10", "task_a"]


def test_spreadsheetbench_system_prompt_matches_skill_rl_baseline(tmp_path: Path) -> None:
    data_root = _make_all_dataset(tmp_path, size=1)

    async def never_called_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
        raise AssertionError("LLM should not be called while reading prompt")

    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=tmp_path / "run", llm=never_called_llm)

    assert env.system_prompt() == SKILL_RL_SPREADSHEET_SYSTEM_PROMPT


class _RecordingEvolutionClient:
    """Evolution client that records the runner's batch lifecycle calls."""

    def __init__(self) -> None:
        self.events: list[str] = []
        self.recorded_cases: list[str] = []
        self.evolve_calls = 0

    async def prepare(self, skill_dirs: list[Path]) -> None:
        self.events.append(f"prepare:{len(skill_dirs)}")

    async def record_case(self, result: Any) -> None:
        self.recorded_cases.append(result.case_id)
        self.events.append(f"record:{result.case_id}")

    async def evolve(self) -> list[Any]:
        self.evolve_calls += 1
        self.events.append("evolve")
        return [
            EvolveOutcome(
                skill_name="xlsx",
                cloud_skill_id="cloud-1",
                evolved=True,
                pending_count=8,
                threshold=4,
                new_version_id=f"v{self.evolve_calls}",
                new_version_ids=[f"v{self.evolve_calls}"],
                summarized_count=2,
                consumed_count=4,
            )
        ]

    async def aclose(self) -> None:
        self.events.append("aclose")


async def _trivial_llm(messages: list[dict[str, Any]], tools: list[dict[str, Any]]) -> dict[str, Any]:
    del tools
    if not any(message.get("role") == "tool" for message in messages):
        return {
            "role": "assistant",
            "content": None,
            "tool_calls": [
                {
                    "id": "call_1",
                    "type": "function",
                    "function": {
                        "name": "shell",
                        "arguments": json.dumps(
                            {
                                "commands": [
                                    "python - <<'PY'\nimport shutil\nshutil.copyfile('input.xlsx','output.xlsx')\nPY"
                                ]
                            }
                        ),
                    },
                }
            ],
        }
    return {"role": "assistant", "content": "Done."}


@pytest.mark.asyncio
async def test_evolution_runs_once_per_batch_of_n_tasks(tmp_path: Path) -> None:
    data_root = _make_all_dataset(tmp_path, size=3)
    client = _RecordingEvolutionClient()
    env = SpreadsheetBenchEnv(
        data_root=data_root,
        run_dir=tmp_path / "run",
        llm=_trivial_llm,
        max_turns=3,
        evolution_client=client,
    )

    result = await env.evaluate(split="all", evolve=True, evolve_every=2, show_progress=False)

    assert result.total == 3
    # 3 cases, evolve_every=2 -> batches [1,2] and [3] -> two evolution passes.
    assert client.evolve_calls == 2
    assert sorted(client.recorded_cases) == ["1", "2", "3"]
    # prepare runs once, before any record/evolve; each batch records before it evolves.
    assert client.events[0] == "prepare:0"
    assert client.events.count("evolve") == 2
    first_evolve = client.events.index("evolve")
    assert client.events[:first_evolve].count("record:1") + client.events[:first_evolve].count("record:2") == 2
    evolution_events = (tmp_path / "run" / "all_evolution_events.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(evolution_events) == 2
    first_event = json.loads(evolution_events[0])
    assert first_event["evolved_at"]
    assert first_event["started_at"]
    assert first_event["ended_at"]
    assert first_event["batch_index"] == 0
    assert first_event["batch_start"] == 0
    assert first_event["batch_end"] == 2
    assert first_event["case_ids"] == ["1", "2"]
    assert first_event["outcomes"][0]["skill_name"] == "xlsx"
    assert first_event["outcomes"][0]["new_version_id"] == "v1"
    assert first_event["error"] is None


@pytest.mark.asyncio
async def test_evolution_batch_progress_callback_fires_per_completed_case(tmp_path: Path) -> None:
    data_root = _make_all_dataset(tmp_path, size=3)
    env = SpreadsheetBenchEnv(data_root=data_root, run_dir=tmp_path / "run", llm=_trivial_llm, max_turns=3)
    delays = {"1": 0.02, "2": 0.01, "3": 0.0}

    async def fake_run_case(case: Any, *, log: bool = True) -> SpreadsheetBenchCaseResult:
        del log
        await asyncio.sleep(delays[case.id])
        return SpreadsheetBenchCaseResult(
            case_id=case.id,
            split=case.split,
            score=1.0,
            finished=True,
            turns=0,
            workdir="",
            messages=[],
        )

    env.run_case = fake_run_case  # type: ignore[method-assign]
    completed: list[str] = []

    results = await env._run_cases_concurrently(
        env.load_cases("all"),
        concurrency=3,
        on_result=lambda result: completed.append(result.case_id),
    )

    assert completed == ["3", "2", "1"]
    assert [result.case_id for result in results] == ["1", "2", "3"]


@pytest.mark.asyncio
async def test_fastapi_evolution_client_registers_records_and_restages(tmp_path: Path) -> None:
    import httpx
    from mindmemos_eval.skills.envs.spreadsheetbench import FastAPISkillEvolutionClient
    from mindmemos_eval.skills.envs.spreadsheetbench.env import SpreadsheetBenchCaseResult
    from mindmemos_sdk.skills import serialize_bundle
    from mindmemos_sdk.transport import HttpTransport

    skill_dir = tmp_path / "skills" / "spreadsheet"
    skill_dir.mkdir(parents=True)
    (skill_dir / "SKILL.md").write_text("original guidance", encoding="utf-8")

    calls: list[tuple[str, str, dict]] = []

    def handler(request: httpx.Request) -> httpx.Response:
        body = json.loads(request.content) if request.content else {}
        path = request.url.path
        calls.append((request.method, path, body))
        if path == "/v1/skills/register":
            return httpx.Response(
                200,
                json={
                    "code": "ok",
                    "data": {
                        "cloud_skill_id": "cloud-1",
                        "version_id": "v1",
                        "content_hash": "hash-v1",
                        "status": "observed",
                    },
                },
            )
        if path == "/v1/memory/add":
            return httpx.Response(200, json={"code": "ok", "data": {"memories": []}})
        if path == "/v1/skills/evolve":
            return httpx.Response(
                200,
                json={
                    "code": "ok",
                    "data": {
                        "cloud_skill_id": "cloud-1",
                        "evolved": True,
                        "pending_count": 4,
                        "threshold": 4,
                        "new_version_id": "v2",
                        "new_version_ids": ["v2"],
                    },
                },
            )
        return httpx.Response(
            200,
            json={
                "code": "ok",
                "data": {
                    "version": {
                        "version_id": "v2",
                        "cloud_skill_id": "cloud-1",
                        "skill_name": "spreadsheet",
                        "content_hash": "hash-v2",
                        "status": "draft",
                        "origin": "cloud",
                        "created_at": "2026-06-17T00:00:00Z",
                    },
                    "content": serialize_bundle({"SKILL.md": "evolved guidance"}),
                },
            },
        )

    transport = HttpTransport(
        base_url="https://api.test", api_key="mk_test", client=httpx.Client(transport=httpx.MockTransport(handler))
    )
    client = FastAPISkillEvolutionClient("https://api.test", api_key="mk_test", transport=transport)

    await client.prepare([skill_dir])
    result = SpreadsheetBenchCaseResult(
        case_id="1",
        split="all",
        score=1.0,
        finished=True,
        turns=2,
        workdir=str(tmp_path),
        messages=[
            {"role": "user", "content": "do the task"},
            {"role": "assistant", "content": None, "tool_calls": [{"function": {"name": "shell", "arguments": "{}"}}]},
            {"role": "assistant", "content": "done"},
        ],
    )
    await client.record_case(result)
    outcomes = await client.evolve()
    await client.aclose()

    # Registered once, recorded the trace with an injected skill_context, evolved, re-staged SKILL.md.
    add_call = next(body for method, path, body in calls if path == "/v1/memory/add")
    assert add_call["mode"] == "async"
    assert add_call["skill_context"][0]["usage"] == "injected"
    assert add_call["skill_context"][0]["base_version_id"] == "v1"
    assert add_call["skill_context"][0]["content_hash"] == "hash-v1"
    assert add_call["score"] == 1.0
    assert add_call["task_id"] == "1"
    assert add_call["metadata"] == {"case_id": "1"}
    assert len(add_call["messages"]) == 3
    evolve_call = next(body for method, path, body in calls if path == "/v1/skills/evolve")
    assert evolve_call["mode"] == "sync"
    assert outcomes[0].evolved is True
    assert outcomes[0].new_version_id == "v2"
    assert (skill_dir / "SKILL.md").read_text(encoding="utf-8") == "evolved guidance"


def test_script_prepare_data_root_extracts_local_archive(tmp_path: Path) -> None:
    source_root = tmp_path / "source"
    _make_all_dataset(source_root, size=1)
    archive = tmp_path / "spreadsheetbench_verified_400.tar.gz"
    with tarfile.open(archive, "w:gz") as tar:
        tar.add(
            source_root / "SpreadsheetBench" / "spreadsheetbench_verified_400",
            arcname="spreadsheetbench_verified_400",
        )

    target_root = tmp_path / "target"
    resolved = prepare_data_root(target_root, archive.as_uri(), download=True)

    assert resolved == target_root
    assert (target_root / "spreadsheetbench_verified_400" / "dataset.json").exists()
