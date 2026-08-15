#!/usr/bin/env python3
"""Transactional formula recalculation for SpreadsheetBench policies."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shlex
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
LIBREOFFICE_PROGRAM_RELATIVE_PATH = Path("libreoffice-appimage/squashfs-root/opt/libreoffice24.2/program")
STATUS_FILENAME = ".tree_only_recalc_status.json"
TASK_LOCAL_HELPER_FILENAME = ".treeskill_recalculate.py"


def stage_task_local_helper(working_dir: str | Path) -> Path:
    """Expose the canonical helper through a stable path inside one task workspace."""

    source = Path(__file__).resolve()
    destination = Path(working_dir).absolute() / TASK_LOCAL_HELPER_FILENAME
    fallback_launcher = f"import runpy\nrunpy.run_path({str(source)!r}, run_name='__main__')\n"
    if destination.is_symlink() or destination.exists():
        if destination.is_symlink() and destination.resolve() == source:
            return destination
        if destination.is_file() and destination.read_text(encoding="utf-8") == fallback_launcher:
            return destination
        raise FileExistsError(f"task-local recalculation helper already exists: {destination}")

    try:
        destination.symlink_to(source)
    except OSError:
        destination.write_text(fallback_launcher, encoding="utf-8")
    return destination


def append_recalculation_instructions(
    prompt: str,
    *,
    working_dir: str | Path,
    output_file: str | Path,
) -> str:
    """Append the canonical transactional recalculation command to a task prompt."""

    workspace = Path(working_dir).absolute()
    helper_path = stage_task_local_helper(workspace)
    output_path = Path(output_file).absolute()
    status_path = workspace / STATUS_FILENAME
    output_argument = (
        output_path.relative_to(workspace).as_posix() if output_path.is_relative_to(workspace) else str(output_path)
    )
    command = shlex.join(
        [
            str(Path(sys.executable).resolve()),
            f"./{helper_path.name}",
            output_argument,
            "--status-path",
            status_path.name,
        ]
    )
    return (
        f"{prompt}\n\n"
        "### formula_recalculation\n"
        "If you create or modify formulas in the output workbook, save it and then "
        "run this exact command before signaling completion:\n"
        f"{command}\n\n"
        "This command is the canonical task-local recalculation command for this run. "
        "Do not substitute another helper or change its paths. A `success` or `not_needed` JSON "
        "status permits completion. If it reports `errors_found` or `failure`, inspect "
        "the reported problem, correct the workbook when possible using task-visible "
        "information, and rerun the same command."
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _file_url(path: Path) -> str:
    return path.resolve().as_uri()


def _local_program_dir() -> Path:
    source = Path(__file__).resolve()
    for parent in source.parents:
        candidate = parent / "tools" / LIBREOFFICE_PROGRAM_RELATIVE_PATH
        if candidate.is_dir():
            return candidate
    return source.parent / "tools" / LIBREOFFICE_PROGRAM_RELATIVE_PATH


def _program_dir_from_soffice_wrapper() -> Path | None:
    executable = shutil.which("soffice")
    if not executable:
        return None
    resolved = Path(executable).resolve()
    if resolved.name != "trace2skill-soffice":
        return None
    return resolved.parent / LIBREOFFICE_PROGRAM_RELATIVE_PATH


def _first_executable(candidates: list[str | Path | None], label: str) -> Path:
    checked: list[str] = []
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate).expanduser().resolve()
        checked.append(str(path))
        if path.is_file() and os.access(path, os.X_OK):
            return path
    raise FileNotFoundError(f"{label} not found or not executable; checked: {checked}")


def _find_libreoffice_programs(
    *,
    soffice_path: str | None = None,
    libreoffice_python: str | None = None,
) -> tuple[Path, Path]:
    local_program_dir = _local_program_dir()
    wrapper_program_dir = _program_dir_from_soffice_wrapper()
    system_soffice = shutil.which("soffice")

    soffice = _first_executable(
        [
            soffice_path,
            os.getenv("TREE_ONLY_SOFFICE_PATH"),
            local_program_dir / "soffice",
            wrapper_program_dir / "soffice" if wrapper_program_dir else None,
            system_soffice,
            "/usr/lib/libreoffice/program/soffice",
            "/opt/libreoffice/program/soffice",
        ],
        "LibreOffice soffice",
    )
    python = _first_executable(
        [
            libreoffice_python,
            os.getenv("TREE_ONLY_LIBREOFFICE_PYTHON"),
            soffice.parent / "python",
            local_program_dir / "python",
            wrapper_program_dir / "python" if wrapper_program_dir else None,
            "/usr/lib/libreoffice/program/python",
            "/opt/libreoffice/program/python",
        ],
        "LibreOffice Python with pyuno",
    )
    return soffice, python


def _workbook_formula_state(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    formulas: list[tuple[str, str]] = []
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append((worksheet.title, cell.coordinate))
    finally:
        workbook.close()

    cached_formula_values = 0
    errors: list[dict[str, str]] = []
    cached = load_workbook(path, data_only=True, read_only=False)
    try:
        for sheet_name, coordinate in formulas:
            value = cached[sheet_name][coordinate].value
            if value is not None:
                cached_formula_values += 1
            if isinstance(value, str):
                for excel_error in EXCEL_ERRORS:
                    if excel_error in value:
                        errors.append(
                            {
                                "sheet": sheet_name,
                                "cell": coordinate,
                                "error": excel_error,
                            }
                        )
                        break
    finally:
        cached.close()

    return {
        "formula_count": len(formulas),
        "cached_formula_value_count": cached_formula_values,
        "excel_errors": errors,
    }


def _write_status(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def recalculate_workbook(
    workbook: Path,
    *,
    status_path: Path | None = None,
    soffice_path: str | None = None,
    libreoffice_python: str | None = None,
    timeout: int = 90,
) -> dict[str, Any]:
    """Recalculate a workbook copy and replace the original only after validation."""

    workbook = workbook.expanduser().resolve()
    status_path = (status_path or workbook.parent / STATUS_FILENAME).expanduser().resolve()
    status: dict[str, Any] = {
        "method": "tree_only_policy_helper",
        "workbook": str(workbook),
        "status": "failure",
        "attempted": False,
        "replaced_output": False,
    }
    if not workbook.is_file():
        status["error"] = f"workbook not found: {workbook}"
        _write_status(status_path, status)
        return status

    work_root: Path | None = None
    office_process: subprocess.Popen[str] | None = None
    office_stdout = ""
    office_stderr = ""

    try:
        before = _workbook_formula_state(workbook)
        status.update(
            {
                "sha256_before": _sha256(workbook),
                "formula_count_before": before["formula_count"],
                "cached_formula_value_count_before": before["cached_formula_value_count"],
            }
        )
        if before["formula_count"] == 0:
            status.update({"status": "not_needed", "error": ""})
            return status

        soffice, office_python = _find_libreoffice_programs(
            soffice_path=soffice_path,
            libreoffice_python=libreoffice_python,
        )
        status.update(
            {
                "attempted": True,
                "soffice": str(soffice),
                "libreoffice_python": str(office_python),
            }
        )

        work_root = Path(tempfile.mkdtemp(prefix=".tree_only_recalc_", dir=workbook.parent))
        profile_dir = work_root / "profile"
        profile_dir.mkdir()
        temporary_workbook = work_root / workbook.name
        shutil.copy2(workbook, temporary_workbook)
        pipe_name = f"tree_only_recalc_{uuid.uuid4().hex}"

        office_command = [
            str(soffice),
            f"-env:UserInstallation={_file_url(profile_dir)}",
            "--headless",
            "--nologo",
            "--nodefault",
            "--norestore",
            "--nofirststartwizard",
            f"--accept=pipe,name={pipe_name};urp;StarOffice.ServiceManager",
        ]
        office_environment = os.environ.copy()
        office_environment["SAL_USE_VCLPLUGIN"] = "svp"
        office_process = subprocess.Popen(
            office_command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            env=office_environment,
        )

        worker_command = [
            str(office_python),
            str(Path(__file__).resolve()),
            "--uno-worker",
            "--pipe-name",
            pipe_name,
            "--timeout",
            str(timeout),
            str(temporary_workbook),
        ]
        worker = subprocess.run(
            worker_command,
            capture_output=True,
            text=True,
            timeout=timeout + 15,
            check=False,
        )
        if worker.returncode != 0:
            detail = worker.stderr.strip() or worker.stdout.strip()
            raise RuntimeError(f"LibreOffice UNO worker failed with exit {worker.returncode}: {detail}")

        after = _workbook_formula_state(temporary_workbook)
        if after["formula_count"] != before["formula_count"]:
            raise RuntimeError(
                f"formula count changed during recalculation: {before['formula_count']} -> {after['formula_count']}"
            )

        os.replace(temporary_workbook, workbook)
        status.update(
            {
                "status": "errors_found" if after["excel_errors"] else "success",
                "error": "",
                "replaced_output": True,
                "sha256_after": _sha256(workbook),
                "formula_count_after": after["formula_count"],
                "cached_formula_value_count_after": after["cached_formula_value_count"],
                "excel_errors": after["excel_errors"],
            }
        )
        return status
    except Exception as exc:
        status["error"] = str(exc)
        return status
    finally:
        if office_process is not None:
            if office_process.poll() is None:
                office_process.terminate()
            try:
                office_stdout, office_stderr = office_process.communicate(timeout=5)
            except subprocess.TimeoutExpired:
                office_process.kill()
                office_stdout, office_stderr = office_process.communicate(timeout=5)
        if status["status"] == "failure":
            status["libreoffice_stdout"] = office_stdout.strip()
            status["libreoffice_stderr"] = office_stderr.strip()
        if work_root is not None:
            shutil.rmtree(work_root, ignore_errors=True)
        _write_status(status_path, status)


def preflight_recalculation_runtime(directory: Path) -> dict[str, Any]:
    """Verify that the configured LibreOffice runtime produces cached formula values."""

    from openpyxl import Workbook, load_workbook

    directory.mkdir(parents=True, exist_ok=True)
    workbook_path = directory / "formula.xlsx"
    status_path = directory / "status.json"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = 1
    worksheet["A2"] = 2
    worksheet["B1"] = "=SUM(A1:A2)"
    workbook.save(workbook_path)
    workbook.close()

    status = recalculate_workbook(workbook_path, status_path=status_path)
    if status["status"] != "success":
        raise RuntimeError(f"transactional recalculation preflight failed: {status.get('error') or status['status']}")
    cached = load_workbook(workbook_path, data_only=True)
    try:
        value = cached.active["B1"].value
    finally:
        cached.close()
    if value != 3:
        raise RuntimeError(f"transactional recalculation preflight returned {value!r}, expected 3")
    return status


def _uno_worker(pipe_name: str, workbook: Path, timeout: int) -> int:
    import uno
    from com.sun.star.beans import PropertyValue

    local_context = uno.getComponentContext()
    resolver = local_context.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver",
        local_context,
    )
    deadline = time.monotonic() + timeout
    context = None
    last_error: Exception | None = None
    while time.monotonic() < deadline:
        try:
            context = resolver.resolve(f"uno:pipe,name={pipe_name};urp;StarOffice.ComponentContext")
            break
        except Exception as exc:
            last_error = exc
            time.sleep(0.2)
    if context is None:
        raise RuntimeError(f"could not connect to LibreOffice UNO pipe: {last_error}")

    service_manager = context.ServiceManager
    desktop = service_manager.createInstanceWithContext("com.sun.star.frame.Desktop", context)
    hidden = PropertyValue()
    hidden.Name = "Hidden"
    hidden.Value = True
    document = desktop.loadComponentFromURL(_file_url(workbook), "_blank", 0, (hidden,))
    if document is None:
        raise RuntimeError(f"LibreOffice could not open workbook: {workbook}")
    try:
        document.calculateAll()
        document.store()
    finally:
        document.close(True)
        desktop.terminate()
    print(json.dumps({"status": "success", "workbook": str(workbook)}))
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Recalculate formulas for the TreeSkill spreadsheet policy")
    parser.add_argument("workbook", nargs="?", type=Path)
    parser.add_argument("--status-path", type=Path, default=None)
    parser.add_argument("--soffice-path", default=None)
    parser.add_argument("--libreoffice-python", default=None)
    parser.add_argument("--timeout", type=int, default=90)
    parser.add_argument("--uno-worker", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--pipe-name", default=None, help=argparse.SUPPRESS)
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    if args.workbook is None:
        raise SystemExit("workbook path is required")
    if args.uno_worker:
        if not args.pipe_name:
            raise SystemExit("--pipe-name is required with --uno-worker")
        return _uno_worker(args.pipe_name, args.workbook.resolve(), args.timeout)

    result = recalculate_workbook(
        args.workbook,
        status_path=args.status_path,
        soffice_path=args.soffice_path,
        libreoffice_python=args.libreoffice_python,
        timeout=args.timeout,
    )
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0 if result["status"] in {"success", "not_needed"} else 1


if __name__ == "__main__":
    sys.exit(main())


__all__ = [
    "STATUS_FILENAME",
    "append_recalculation_instructions",
    "preflight_recalculation_runtime",
    "recalculate_workbook",
]
