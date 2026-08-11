"""Official cached-cell-value SpreadsheetBench comparison semantics."""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

from ....errors import SkillCapabilityUnavailableError


def _openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise SkillCapabilityUnavailableError(
            "SpreadsheetBench requires the mindmemos-skill[spreadsheetbench] optional dependency"
        ) from exc
    return openpyxl


def _transform_value(value: Any) -> Any:
    if isinstance(value, bool):
        return round(float(value), 2)
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, datetime.time):
        return str(value)[:-3]
    if isinstance(value, datetime.datetime):
        delta = value - datetime.datetime(1899, 12, 30)
        return round(delta.days + delta.seconds / 86400.0, 0)
    if isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            return value
    return value


def _same_value(left: Any, right: Any) -> bool:
    left, right = _transform_value(left), _transform_value(right)
    if left in {"", None} and right in {"", None}:
        return True
    return type(left) is type(right) and left == right


def _cells(cell_range: str) -> list[str]:
    if ":" not in cell_range:
        return [cell_range]
    openpyxl = _openpyxl()
    start, end = cell_range.split(":", 1)
    start_column = openpyxl.utils.column_index_from_string("".join(filter(str.isalpha, start)))
    end_column = openpyxl.utils.column_index_from_string("".join(filter(str.isalpha, end)))
    start_row = int("".join(filter(str.isdigit, start)))
    end_row = int("".join(filter(str.isdigit, end)))
    return [
        f"{openpyxl.utils.get_column_letter(column)}{row}"
        for column in range(start_column, end_column + 1)
        for row in range(start_row, end_row + 1)
    ]


def compare_workbooks(
    golden_path: str | Path,
    output_path: str | Path,
    answer_position: str,
) -> tuple[bool, str]:
    output_path = Path(output_path)
    if not output_path.exists():
        return False, "file not exist"
    openpyxl = _openpyxl()
    try:
        golden = openpyxl.load_workbook(golden_path, data_only=True)
        output = openpyxl.load_workbook(output_path, data_only=True)
    except Exception as exc:
        return False, f"load error: {exc}"
    try:
        for target in filter(None, (part.strip() for part in answer_position.split(","))):
            if "!" in target:
                sheet, cell_range = target.split("!", 1)
                sheet = sheet.strip().strip("'\"")
            else:
                sheet, cell_range = golden.sheetnames[0], target
            cell_range = cell_range.strip().strip("'\"")
            if sheet not in output.sheetnames:
                return False, f"worksheet not found: {sheet}"
            for cell in _cells(cell_range):
                expected = golden[sheet][cell].value
                actual = output[sheet][cell].value
                if not _same_value(expected, actual):
                    return False, f"value@{sheet}!{cell}: gt={expected!r} pred={actual!r}"
        return True, ""
    finally:
        golden.close()
        output.close()


__all__ = ["compare_workbooks"]
