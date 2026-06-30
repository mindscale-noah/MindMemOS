"""SpreadsheetBench workbook value comparator."""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any


def _load_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised only when dependency is missing
        raise RuntimeError("SpreadsheetBench evaluation requires openpyxl. Install it with `uv add openpyxl`.") from exc
    return openpyxl


def _datetime_to_float(value: _dt.datetime) -> float:
    excel_start_date = _dt.datetime(1899, 12, 30)
    delta = value - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(value: Any) -> Any:
    """Apply SpreadsheetBench's official value normalization."""
    if isinstance(value, bool):
        return round(float(value), 2)
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, _dt.time):
        return str(value)[:-3]
    if isinstance(value, _dt.datetime):
        return round(_datetime_to_float(value), 0)
    if isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            return value
    return value


def compare_cell_value(left: Any, right: Any) -> bool:
    """Compare two cell values with SpreadsheetBench semantics."""
    left = transform_value(left)
    right = transform_value(right)
    if (left == "" and right is None) or (left is None and right == ""):
        return True
    if (left == "" and right == "") or (left is None and right is None):
        return True
    if type(left) is not type(right):
        return False
    return left == right


def _col_num_to_name(number: int) -> str:
    name = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _col_name_to_num(name: str) -> int:
    number = 0
    for char in name:
        number = number * 26 + (ord(char.upper()) - ord("A") + 1)
    return number


def _parse_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    start_cell, end_cell = range_str.split(":")
    start_col = "".join(char for char in start_cell if char.isalpha())
    start_row = "".join(char for char in start_cell if char.isdigit())
    end_col = "".join(char for char in end_cell if char.isalpha())
    end_row = "".join(char for char in end_cell if char.isdigit())
    return (_col_name_to_num(start_col), int(start_row)), (_col_name_to_num(end_col), int(end_row))


def generate_cell_names(range_str: str) -> list[str]:
    """Expand an Excel cell or range string into individual cell names."""
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = _parse_range(range_str)
    cols = [_col_num_to_name(index) for index in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in cols for row in range(start_row, end_row + 1)]


def _cell_level_compare(wb_gt: Any, wb_proc: Any, sheet_name: str, cell_range: str) -> tuple[bool, str]:
    if sheet_name not in wb_proc.sheetnames:
        return False, f"worksheet not found: {sheet_name}"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]
    for cell_name in generate_cell_names(cell_range):
        gt_cell = ws_gt[cell_name]
        proc_cell = ws_proc[cell_name]
        if not compare_cell_value(gt_cell.value, proc_cell.value):
            return False, f"value@{sheet_name}!{cell_name}: gt={gt_cell.value!r} pred={proc_cell.value!r}"
    return True, ""


def compare_workbooks(gt_file: str | Path, proc_file: str | Path, answer_position: str) -> tuple[bool, str]:
    """Return ``(ok, message)`` for a golden/produced workbook comparison."""
    gt_path = Path(gt_file)
    proc_path = Path(proc_file)
    if not proc_path.exists():
        return False, "file not exist"

    openpyxl = _load_openpyxl()
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_path, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - normalize workbook load errors for result records
        return False, f"load error: {exc}"
    try:
        ok_all = True
        first_message = ""
        for sheet_cell_range in (answer_position or "").split(","):
            sheet_cell_range = sheet_cell_range.strip()
            if not sheet_cell_range:
                continue
            if "!" in sheet_cell_range:
                sheet_name, cell_range = sheet_cell_range.split("!", 1)
                sheet_name = sheet_name.strip().strip("'\"")
            else:
                sheet_name = wb_gt.sheetnames[0]
                cell_range = sheet_cell_range
            ok, message = _cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range.strip().strip("'\""))
            if not ok:
                ok_all = False
                first_message = first_message or message
        return ok_all, first_message
    finally:
        wb_gt.close()
        wb_proc.close()
