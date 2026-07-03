"""Offline metrics collector for benchmark matrix runs."""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import statistics
import urllib.error
import urllib.parse
import urllib.request
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Protocol

from qdrant_client import AsyncQdrantClient
from qdrant_client import models as qmodels

DEFAULT_QDRANT_URL = "http://localhost:6333"
DEFAULT_CLICKHOUSE_URL = "http://localhost:8123"
DEFAULT_CLICKHOUSE_DATABASE = "otel"
DEFAULT_CLICKHOUSE_TABLE = "otel_traces"
DEFAULT_CLICKHOUSE_USER = "mindmemos"
DEFAULT_CLICKHOUSE_PASSWORD = "mindmemos_dev_password"
DEFAULT_MEMORY_COLLECTION = "memory_item_v1"
DEFAULT_ADD_RECORD_COLLECTION = "add_record_v1"
DEFAULT_SEARCH_RECORD_COLLECTION = "search_record_v1"


class QdrantScrollClient(Protocol):
    """Minimal Qdrant client surface used by the collector."""

    async def scroll(
        self,
        collection_name: str,
        *,
        scroll_filter: qmodels.Filter | None = None,
        limit: int = 100,
        offset: Any | None = None,
        with_payload: bool = True,
        with_vectors: bool = False,
    ) -> tuple[list[Any], Any | None]:
        """Scroll points from one collection."""
        ...


@dataclass(frozen=True)
class QdrantCollectorConfig:
    """Qdrant collection names and scan settings."""

    url: str = DEFAULT_QDRANT_URL
    api_key: str | None = None
    memory_collection: str = DEFAULT_MEMORY_COLLECTION
    add_record_collection: str = DEFAULT_ADD_RECORD_COLLECTION
    search_record_collection: str = DEFAULT_SEARCH_RECORD_COLLECTION
    page_size: int = 256


@dataclass(frozen=True)
class ClickHouseCollectorConfig:
    """ClickHouse connection and table settings."""

    url: str = DEFAULT_CLICKHOUSE_URL
    database: str = DEFAULT_CLICKHOUSE_DATABASE
    table: str = DEFAULT_CLICKHOUSE_TABLE
    user: str = DEFAULT_CLICKHOUSE_USER
    password: str = DEFAULT_CLICKHOUSE_PASSWORD
    timeout_seconds: float = 10.0
    enabled: bool = True
    strict: bool = False


@dataclass(frozen=True)
class MetricsCollectorConfig:
    """Offline collector configuration."""

    qdrant: QdrantCollectorConfig = field(default_factory=QdrantCollectorConfig)
    clickhouse: ClickHouseCollectorConfig = field(default_factory=ClickHouseCollectorConfig)


@dataclass(frozen=True)
class BenchmarkRunRef:
    """Identity fields read from one benchmark manifest row."""

    benchmark: str
    run_id: str
    project_id: str
    key_id: str
    memory_algorithm: str
    request_ids: dict[str, list[str]]
    raw: dict[str, Any] = field(default_factory=dict)

    @classmethod
    def from_manifest(cls, row: dict[str, Any]) -> "BenchmarkRunRef":
        """Create a run reference from a manifest JSON object."""
        return cls(
            benchmark=str(row.get("benchmark") or ""),
            run_id=str(row.get("run_id") or ""),
            project_id=str(row.get("project_id") or ""),
            key_id=str(row.get("key_id") or ""),
            memory_algorithm=str(row.get("memory_algorithm") or ""),
            request_ids={stage: list(ids or []) for stage, ids in (row.get("request_ids") or {}).items()},
            raw=dict(row),
        )


def load_manifest(path: str | Path) -> list[BenchmarkRunRef]:
    """Load benchmark manifest JSONL rows."""
    rows: list[BenchmarkRunRef] = []
    for line_no, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"invalid manifest JSON on line {line_no}: {exc}") from exc
        rows.append(BenchmarkRunRef.from_manifest(payload))
    return rows


async def collect_benchmark_metrics(
    manifest_path: str | Path,
    output_path: str | Path,
    config: MetricsCollectorConfig,
    table_output_path: str | Path | None = None,
    xlsx_output_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    """Collect Qdrant and ClickHouse metrics for every manifest row."""
    runs = load_manifest(manifest_path)
    rows: list[dict[str, Any]] = []
    qdrant = AsyncQdrantClient(url=config.qdrant.url, api_key=config.qdrant.api_key, trust_env=False)
    try:
        for run in runs:
            qdrant_metrics = await collect_qdrant_metrics(qdrant, run, config.qdrant)
            clickhouse_metrics = await collect_clickhouse_metrics(run, config.clickhouse)
            row = {
                **run.raw,
                "metrics": {
                    "qdrant": qdrant_metrics,
                    "clickhouse": clickhouse_metrics,
                },
                "metrics_collected_at": datetime.now(UTC).isoformat(),
            }
            rows.append(row)
    finally:
        await qdrant.close()

    write_jsonl(output_path, rows)
    if table_output_path:
        write_metrics_table(table_output_path, rows)
    if xlsx_output_path:
        write_metrics_xlsx(xlsx_output_path, rows)
    return rows


async def collect_qdrant_metrics(
    client: QdrantScrollClient,
    run: BenchmarkRunRef,
    config: QdrantCollectorConfig | None = None,
) -> dict[str, Any]:
    """Collect add/search timings and memory counts from Qdrant."""
    cfg = config or QdrantCollectorConfig()
    add_records = await _scroll_payloads(
        client,
        cfg.add_record_collection,
        _identity_filter(run.project_id, run.key_id, _request_ids(run, "add")),
        page_size=cfg.page_size,
    )
    search_records = await _scroll_payloads(
        client,
        cfg.search_record_collection,
        _identity_filter(run.project_id, run.key_id, _request_ids(run, "search")),
        page_size=cfg.page_size,
    )
    memory_records = await _scroll_payloads(
        client,
        cfg.memory_collection,
        _identity_filter(run.project_id, run.key_id, []),
        page_size=cfg.page_size,
    )

    add_durations = [d for p in add_records if (d := _duration_ms(p)) is not None]
    search_durations = [d for p in search_records if (d := _duration_ms(p)) is not None]

    return {
        "add_record_count": len(add_records),
        "search_record_count": len(search_records),
        "memory_count": len(memory_records),
        "add_record_memory_count": sum(_payload_memory_count(payload) for payload in add_records),
        "add_total_time_ms": _total_duration_ms(add_records),
        "search_total_time_ms": _total_duration_ms(search_records),
        "add_avg_time_ms": _avg_duration_ms(add_records),
        "search_avg_time_ms": _avg_duration_ms(search_records),
        "add_percentiles_ms": _percentile_stats(add_durations),
        "search_percentiles_ms": _percentile_stats(search_durations),
    }


async def collect_clickhouse_metrics(run: BenchmarkRunRef, config: ClickHouseCollectorConfig) -> dict[str, Any]:
    """Collect LLM span usage metrics from ClickHouse."""
    if not config.enabled:
        return {"enabled": False}
    request_ids = _all_request_ids(run)
    if not request_ids:
        return {"enabled": True, "llm_call_count": 0, "llm_by_task": []}

    try:
        totals = await asyncio.to_thread(_query_clickhouse_json, config, _llm_totals_sql(run, request_ids, config))
        by_task = await asyncio.to_thread(_query_clickhouse_json, config, _llm_by_task_sql(run, request_ids, config))
        token_percentiles = await asyncio.to_thread(
            _query_clickhouse_json, config, _llm_token_percentiles_sql(run, request_ids, config)
        )
        search_token_perc = await asyncio.to_thread(
            _query_clickhouse_json, config, _search_token_percentiles_sql(run, request_ids, config)
        )
    except Exception as exc:  # noqa: BLE001 - collector should keep producing Qdrant metrics.
        if config.strict:
            raise
        return {"enabled": True, "error": str(exc), "llm_call_count": 0, "llm_by_task": []}

    total_row = (totals.get("data") or [{}])[0]
    token_perc_data = token_percentiles.get("data") or []
    token_perc = {item.get("task"): item for item in token_perc_data if isinstance(item, dict)}
    search_token_perc_row = (search_token_perc.get("data") or [{}])[0]

    return {
        "enabled": True,
        "llm_call_count": _to_int(total_row.get("llm_call_count")),
        "llm_total_tokens": _to_int(total_row.get("llm_total_tokens")),
        "llm_prompt_tokens": _to_int(total_row.get("llm_prompt_tokens")),
        "llm_completion_tokens": _to_int(total_row.get("llm_completion_tokens")),
        "llm_total_time_ms": _to_float(total_row.get("llm_total_time_ms")),
        "llm_by_task": by_task.get("data") or [],
        "llm_token_percentiles_by_task": token_perc,
        "search_token_percentiles": search_token_perc_row,
    }


async def _scroll_payloads(
    client: QdrantScrollClient,
    collection_name: str,
    scroll_filter: qmodels.Filter,
    *,
    page_size: int,
) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    cursor: Any | None = None
    while True:
        records, cursor = await client.scroll(
            collection_name=collection_name,
            scroll_filter=scroll_filter,
            limit=page_size,
            offset=cursor,
            with_payload=True,
            with_vectors=False,
        )
        payloads.extend(dict(getattr(record, "payload", None) or {}) for record in records)
        if cursor is None:
            return payloads


def _identity_filter(project_id: str, key_id: str, request_ids: Sequence[str]) -> qmodels.Filter:
    must: list[Any] = [
        qmodels.FieldCondition(key="project_id", match=qmodels.MatchValue(value=project_id)),
        qmodels.FieldCondition(key="api_key_uuid", match=qmodels.MatchValue(value=key_id)),
    ]
    if request_ids:
        must.append(qmodels.FieldCondition(key="request_id", match=qmodels.MatchAny(any=list(request_ids))))
    return qmodels.Filter(must=must)


def _request_ids(run: BenchmarkRunRef, stage: str) -> list[str]:
    return [str(value) for value in run.request_ids.get(stage, []) if value]


def _all_request_ids(run: BenchmarkRunRef) -> list[str]:
    ids: list[str] = []
    for stage in ("add", "search", "answer", "eval"):
        ids.extend(_request_ids(run, stage))
    return ids


def _payload_memory_count(payload: dict[str, Any]) -> int:
    memories = payload.get("memories")
    return len(memories) if isinstance(memories, list) else 0


def _duration_ms(payload: dict[str, Any]) -> float | None:
    started = _parse_dt(payload.get("request_submitted_at"))
    finished = _parse_dt(payload.get("task_completed_at"))
    if started is None or finished is None:
        return None
    return max((finished - started).total_seconds() * 1000.0, 0.0)


def _total_duration_ms(payloads: Iterable[dict[str, Any]]) -> float:
    return round(sum(value for payload in payloads if (value := _duration_ms(payload)) is not None), 3)


def _avg_duration_ms(payloads: Sequence[dict[str, Any]]) -> float | None:
    durations = [value for payload in payloads if (value := _duration_ms(payload)) is not None]
    if not durations:
        return None
    return round(sum(durations) / len(durations), 3)


def _aggregate_token_percentiles(token_perc: dict[str, Any], op_name: str) -> dict[str, int | None]:
    """Aggregate token percentiles for all tasks with op_name (e.g., 'memory.add.*', 'search.*').

    For each stat (min/max/p50/p95), collects the values across all matching tasks
    and computes the median and p95 to estimate the operation-level distribution.

    Args:
        token_perc: Dict mapping task name to token percentile dict (from ClickHouse).
        op_name: Operation name ('add' or 'search').

    Returns:
        Dict with 'min', 'max', 'p50', 'p95' keys (or None if no data).
    """
    token_mins, token_maxs, token_p50s, token_p95s = [], [], [], []

    for task_name, perc in token_perc.items():
        if not isinstance(task_name, str) or not isinstance(perc, dict):
            continue
        # Match 'memory.add.*' for op_name='add', or 'search.*' for op_name='search'
        if op_name == "add":
            matches = task_name.startswith("memory.add.")
        elif op_name == "search":
            matches = task_name.startswith("search.")
        else:
            continue
        if not matches:
            continue
        # Collect per-task percentile values
        if (v := perc.get("token_min")) is not None:
            token_mins.append(_to_int(v))
        if (v := perc.get("token_max")) is not None:
            token_maxs.append(_to_int(v))
        if (v := perc.get("token_p50")) is not None:
            token_p50s.append(_to_int(v))
        if (v := perc.get("token_p95")) is not None:
            token_p95s.append(_to_int(v))

    if not (token_mins or token_maxs or token_p50s or token_p95s):
        return {"min": None, "max": None, "p50": None, "p95": None}

    # Compute min of all task mins, max of all task maxes, etc.
    return {
        "min": min(token_mins) if token_mins else None,
        "max": max(token_maxs) if token_maxs else None,
        "p50": int(statistics.median(token_p50s)) if token_p50s else None,
        "p95": int(_percentile(sorted(token_p95s), 95)) if token_p95s else None,
    }


def _task_prefix_avg_tokens(by_task: list[Any], prefix: str) -> float | None:
    """Average tokens per LLM call across all tasks matching ``prefix`` (e.g. 'memory.add.')."""
    total_tokens = 0
    call_count = 0
    for task in by_task:
        if not isinstance(task, dict) or not str(task.get("task") or "").startswith(prefix):
            continue
        total_tokens += _to_int(task.get("llm_total_tokens"))
        call_count += _to_int(task.get("llm_call_count"))
    return round(total_tokens / call_count, 1) if call_count else None


def _eval_qa_token_series(row: dict[str, Any]) -> dict[str, list[float]]:
    """Extract per-question token counts from manifest eval_result for percentile computation.

    Returns a dict mapping stage name ('answer', 'judge') to a list of total_tokens
    values, one per question.  The manifest stores full qa_results for locomo and
    longmemeval; personamem uses deterministic scoring so answer/judge are absent.
    """
    er = row.get("eval_result") or {}
    benchmark = row.get("benchmark", "")
    series: dict[str, list[float]] = {"answer": [], "judge": []}

    if benchmark == "locomo":
        qa_lists = [
            conv.get("qa_results") or []
            for conv in (er.get("conversations") or [])
        ]
    elif benchmark == "longmemeval":
        qa_lists = [
            s.get("qa_results") or []
            for s in (er.get("samples") or [])
        ]
    else:
        return series

    for qa_list in qa_lists:
        for qa in qa_list:
            if not isinstance(qa, dict):
                continue
            for stage in ("answer", "judge"):
                v = qa.get(f"{stage}_total_tokens")
                if v is not None:
                    series[stage].append(float(v))

    return series


def _percentile_stats(values: Sequence[float]) -> dict[str, float | None]:
    """Compute min, max, p50, p95 percentiles for a list of values."""
    if not values:
        return {"min": None, "max": None, "p50": None, "p95": None}
    sorted_vals = sorted(values)
    return {
        "min": round(sorted_vals[0], 3),
        "max": round(sorted_vals[-1], 3),
        "p50": round(statistics.median(sorted_vals), 3),
        "p95": round(_percentile(sorted_vals, 95), 3),
    }


def _percentile(values: Sequence[float], p: float) -> float:
    """Compute p-th percentile (0-100) of sorted values using linear interpolation."""
    if not values:
        return 0.0
    if p <= 0:
        return values[0]
    if p >= 100:
        return values[-1]
    h = (len(values) - 1) * p / 100.0
    i = int(h)
    frac = h - i
    if i + 1 < len(values):
        return values[i] * (1 - frac) + values[i + 1] * frac
    return values[i]


def _parse_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


def _llm_totals_sql(run: BenchmarkRunRef, request_ids: Sequence[str], config: ClickHouseCollectorConfig) -> str:
    del request_ids  # project_id + api_key_uuid already scope one benchmark run uniquely.
    table = _qualified_table(config)
    return f"""
WITH benchmark_traces AS (
    SELECT DISTINCT TraceId
    FROM {table}
    WHERE SpanAttributes['project_id'] = {_sql_literal(run.project_id)}
      AND SpanAttributes['api_key_uuid'] = {_sql_literal(run.key_id)}
)
SELECT
    count() AS llm_call_count,
    sum(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])) AS llm_total_tokens,
    sum(toInt64OrZero(SpanAttributes['llm.usage.prompt_tokens'])) AS llm_prompt_tokens,
    sum(toInt64OrZero(SpanAttributes['llm.usage.completion_tokens'])) AS llm_completion_tokens,
    round(sum(Duration) / 1000000, 3) AS llm_total_time_ms
FROM {table}
WHERE SpanName = 'llm.chat'
  AND TraceId IN (SELECT TraceId FROM benchmark_traces)
FORMAT JSON
""".strip()


def _llm_by_task_sql(run: BenchmarkRunRef, request_ids: Sequence[str], config: ClickHouseCollectorConfig) -> str:
    del request_ids  # project_id + api_key_uuid already scope one benchmark run uniquely.
    table = _qualified_table(config)
    return f"""
WITH benchmark_traces AS (
    SELECT DISTINCT TraceId
    FROM {table}
    WHERE SpanAttributes['project_id'] = {_sql_literal(run.project_id)}
      AND SpanAttributes['api_key_uuid'] = {_sql_literal(run.key_id)}
)
SELECT
    SpanAttributes['llm.task'] AS task,
    count() AS llm_call_count,
    sum(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])) AS llm_total_tokens,
    sum(toInt64OrZero(SpanAttributes['llm.usage.prompt_tokens'])) AS llm_prompt_tokens,
    sum(toInt64OrZero(SpanAttributes['llm.usage.completion_tokens'])) AS llm_completion_tokens,
    round(sum(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])) / count(), 3) AS llm_avg_tokens,
    round(sum(Duration) / 1000000, 3) AS llm_total_time_ms
FROM {table}
WHERE SpanName = 'llm.chat'
  AND TraceId IN (SELECT TraceId FROM benchmark_traces)
GROUP BY task
ORDER BY task
FORMAT JSON
""".strip()


def _llm_token_percentiles_sql(
    run: BenchmarkRunRef, request_ids: Sequence[str], config: ClickHouseCollectorConfig
) -> str:
    """Compute token count percentiles per LLM task."""
    del request_ids  # project_id + api_key_uuid already scope one benchmark run uniquely.
    table = _qualified_table(config)
    return f"""
WITH benchmark_traces AS (
    SELECT DISTINCT TraceId
    FROM {table}
    WHERE SpanAttributes['project_id'] = {_sql_literal(run.project_id)}
      AND SpanAttributes['api_key_uuid'] = {_sql_literal(run.key_id)}
)
SELECT
    SpanAttributes['llm.task'] AS task,
    min(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])) AS token_min,
    max(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])) AS token_max,
    round(quantile(0.5)(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])), 0) AS token_p50,
    round(quantile(0.95)(toInt64OrZero(SpanAttributes['llm.usage.total_tokens'])), 0) AS token_p95
FROM {table}
WHERE SpanName = 'llm.chat'
  AND TraceId IN (SELECT TraceId FROM benchmark_traces)
GROUP BY task
ORDER BY task
FORMAT JSON
""".strip()


def _search_token_percentiles_sql(
    run: BenchmarkRunRef, request_ids: Sequence[str], config: ClickHouseCollectorConfig
) -> str:
    """Compute search-stage token totals and per-query percentiles, entirely from ClickHouse.

    Groups by TraceId so that all search.* LLM calls within one search request
    are summed before computing percentiles — giving per-query cost distribution
    rather than per-LLM-call distribution. Also returns call/query counts and
    prompt/completion/total token sums so no Qdrant data is needed for search
    token summaries.
    """
    del request_ids  # project_id + api_key_uuid already scope one benchmark run uniquely.
    table = _qualified_table(config)
    return f"""
WITH benchmark_traces AS (
    SELECT DISTINCT TraceId
    FROM {table}
    WHERE SpanAttributes['project_id'] = {_sql_literal(run.project_id)}
      AND SpanAttributes['api_key_uuid'] = {_sql_literal(run.key_id)}
),
search_calls AS (
    SELECT
        TraceId,
        toInt64OrZero(SpanAttributes['llm.usage.total_tokens']) AS total_tokens,
        toInt64OrZero(SpanAttributes['llm.usage.prompt_tokens']) AS prompt_tokens,
        toInt64OrZero(SpanAttributes['llm.usage.completion_tokens']) AS completion_tokens
    FROM {table}
    WHERE SpanName = 'llm.chat'
      AND startsWith(SpanAttributes['llm.task'], 'search.')
      AND TraceId IN (SELECT TraceId FROM benchmark_traces)
),
per_query AS (
    SELECT
        TraceId,
        sum(total_tokens) AS total_tokens
    FROM search_calls
    GROUP BY TraceId
)
SELECT
    (SELECT count() FROM search_calls)              AS call_count,
    (SELECT sum(prompt_tokens) FROM search_calls)    AS prompt_tokens,
    (SELECT sum(completion_tokens) FROM search_calls) AS completion_tokens,
    (SELECT sum(total_tokens) FROM search_calls)     AS total_tokens,
    count()                                          AS query_count,
    min(total_tokens)                                AS token_min,
    max(total_tokens)                                AS token_max,
    round(avg(total_tokens), 1)                      AS token_avg,
    round(quantile(0.5)(total_tokens), 0)            AS token_p50,
    round(quantile(0.95)(total_tokens), 0)           AS token_p95
FROM per_query
FORMAT JSON
""".strip()


def _query_clickhouse_json(config: ClickHouseCollectorConfig, sql: str) -> dict[str, Any]:
    params = urllib.parse.urlencode(
        {
            "user": config.user,
            "password": config.password,
            "database": config.database,
        }
    )
    url = f"{config.url.rstrip('/')}?{params}"
    body = sql.encode("utf-8")
    request = urllib.request.Request(url, data=body, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=config.timeout_seconds) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"clickhouse query failed: HTTP {exc.code}: {detail}") from exc


def _qualified_table(config: ClickHouseCollectorConfig) -> str:
    return f"{_sql_identifier(config.database)}.{_sql_identifier(config.table)}"


def _sql_literal(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def _sql_identifier(value: str) -> str:
    if not value.replace("_", "").isalnum():
        raise ValueError(f"unsafe ClickHouse identifier: {value!r}")
    return f"`{value}`"


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _to_float(value: Any) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def write_jsonl(path: str | Path, rows: Iterable[dict[str, Any]]) -> None:
    """Write rows as JSONL."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")


SUMMARY_HEADERS = [
    "benchmark",
    "run_id",
    "key_id",
    "project_id",
    "algorithm",
    "started_at",
    "finished_at",
    "add_count",
    "search_count",
    "memory_count",
    "add_total_ms",
    "search_total_ms",
    "add_avg_ms",
    "search_avg_ms",
    "llm_calls",
    "llm_total_tokens",
    "llm_prompt_tokens",
    "llm_completion_tokens",
    "llm_total_time_ms",
    "search_total_tokens",
    "search_avg_tokens/query",
    "search_avg_tokens/call",
    "overall_accuracy",
    "answer_llm_calls",
    "answer_prompt_tokens",
    "answer_completion_tokens",
    "answer_total_tokens",
    "judge_llm_calls",
    "judge_prompt_tokens",
    "judge_completion_tokens",
    "judge_total_tokens",
    "add_token_avg",
    "answer_token_avg",
    "judge_token_avg",
]

EVAL_METRICS_HEADERS = [
    "benchmark",
    "run_id",
    "protocol",
    "official_protocol_commit",
    "benchmark_version",
    "context_size",
    "evaluation_mode",
    "overall_accuracy",
    "correct",
    "total",
    "scope_total",
    "scope_build_success",
    "scope_build_failure",
    "scope_violation_count",
    "search_failure_count",
    "answer_failure_count",
    "search_llm_calls",
    "search_prompt_tokens",
    "search_completion_tokens",
    "search_total_tokens",
    "answer_llm_calls",
    "answer_prompt_tokens",
    "answer_completion_tokens",
    "answer_total_tokens",
    "judge_llm_calls",
    "judge_prompt_tokens",
    "judge_completion_tokens",
    "judge_total_tokens",
    "build_elapsed_seconds",
    "search_elapsed_seconds",
    "answer_elapsed_seconds",
    "total_elapsed_seconds",
    "by_question_type",
    "by_topic",
]


def _summary_values(row: dict[str, Any]) -> list[Any]:
    """Project one metrics row onto ``SUMMARY_HEADERS`` order.

    ``memory_count`` is the only memory tally surfaced: the final number of points
    persisted in ``memory_item_v1`` for the run. The per-request ``add_memories``
    echo is intentionally omitted (it under-counts derived profile/fact memories).
    """
    qdrant = dict(row.get("metrics", {}).get("qdrant") or {})
    clickhouse = dict(row.get("metrics", {}).get("clickhouse") or {})
    sp = clickhouse.get("search_token_percentiles") or {}
    em = dict((row.get("eval_result") or {}).get("metrics") or {})
    search_call_count = _to_int(sp.get("call_count"))
    search_total_tokens = _to_int(sp.get("total_tokens")) if sp.get("total_tokens") is not None else None
    search_avg_per_query = _to_float(sp["token_avg"]) if sp.get("token_avg") is not None else None
    search_avg_per_call = (
        round(search_total_tokens / search_call_count, 3) if search_total_tokens and search_call_count else None
    )
    values: list[Any] = [
        row.get("benchmark"),
        row.get("run_id"),
        row.get("key_id"),
        row.get("project_id"),
        row.get("memory_algorithm"),
        row.get("started_at"),
        row.get("finished_at"),
        qdrant.get("add_record_count"),
        qdrant.get("search_record_count"),
        qdrant.get("memory_count"),
        qdrant.get("add_total_time_ms"),
        qdrant.get("search_total_time_ms"),
        qdrant.get("add_avg_time_ms"),
        qdrant.get("search_avg_time_ms"),
        clickhouse.get("llm_call_count"),
        clickhouse.get("llm_total_tokens"),
        clickhouse.get("llm_prompt_tokens"),
        clickhouse.get("llm_completion_tokens"),
        clickhouse.get("llm_total_time_ms"),
        search_total_tokens,
        search_avg_per_query,
        search_avg_per_call,
        em.get("overall_accuracy"),
        em.get("answer_llm_calls"),
        em.get("answer_prompt_tokens"),
        em.get("answer_completion_tokens"),
        em.get("answer_total_tokens"),
        em.get("judge_llm_calls"),
        em.get("judge_prompt_tokens"),
        em.get("judge_completion_tokens"),
        em.get("judge_total_tokens"),
    ]
    add_avg = _task_prefix_avg_tokens(clickhouse.get("llm_by_task") or [], "memory.add.")
    qa_series = _eval_qa_token_series(row)
    ans_vals = qa_series.get("answer") or []
    jdg_vals = qa_series.get("judge") or []
    ans_avg = round(sum(ans_vals) / len(ans_vals), 1) if ans_vals else None
    jdg_avg = round(sum(jdg_vals) / len(jdg_vals), 1) if jdg_vals else None
    values += [add_avg, ans_avg, jdg_avg]
    return values


def _eval_metrics_values(row: dict[str, Any]) -> list[Any]:
    """Project benchmark-owned evaluation metrics without changing collector semantics."""
    import warnings

    eval_result = dict(row.get("eval_result") or {})
    metrics = dict(eval_result.get("metrics") or {})
    sp = dict(row.get("metrics", {}).get("clickhouse", {}).get("search_token_percentiles") or {})

    # ✅ 检查关键字段，防止静默数据丢失
    if eval_result and not eval_result.get("protocol"):
        warnings.warn(
            f"eval_result for run_id={row.get('run_id')} missing 'protocol' field",
            UserWarning,
            stacklevel=2,
        )
    if metrics and not metrics.get("overall_accuracy"):
        warnings.warn(
            f"eval_result metrics for run_id={row.get('run_id')} missing 'overall_accuracy'",
            UserWarning,
            stacklevel=2,
        )
    return [
        row.get("benchmark"),
        row.get("run_id"),
        eval_result.get("protocol"),
        eval_result.get("official_protocol_commit"),
        eval_result.get("benchmark_version"),
        eval_result.get("context_size"),
        eval_result.get("evaluation_mode"),
        metrics.get("overall_accuracy"),
        metrics.get("correct"),
        metrics.get("total_questions") or metrics.get("total"),
        metrics.get("scope_total"),
        metrics.get("scope_build_success"),
        metrics.get("scope_build_failure"),
        metrics.get("scope_violation_count"),
        metrics.get("search_failure_count"),
        metrics.get("answer_failure_count"),
        metrics.get("search_llm_calls") or _to_int(sp.get("call_count")),
        metrics.get("search_prompt_tokens") or _to_int(sp.get("prompt_tokens")),
        metrics.get("search_completion_tokens") or _to_int(sp.get("completion_tokens")),
        metrics.get("search_total_tokens") or _to_int(sp.get("total_tokens")),
        metrics.get("answer_llm_calls"),
        metrics.get("answer_prompt_tokens"),
        metrics.get("answer_completion_tokens"),
        metrics.get("answer_total_tokens"),
        metrics.get("judge_llm_calls"),
        metrics.get("judge_prompt_tokens"),
        metrics.get("judge_completion_tokens"),
        metrics.get("judge_total_tokens"),
        metrics.get("build_elapsed_seconds"),
        metrics.get("search_elapsed_seconds"),
        metrics.get("answer_elapsed_seconds"),
        metrics.get("total_elapsed_seconds"),
        metrics.get("by_question_type"),
        metrics.get("by_topic"),
    ]


def write_metrics_table(path: str | Path, rows: Iterable[dict[str, Any]]) -> None:
    """Write a compact Markdown table for human review."""
    lines = [
        "| " + " | ".join(SUMMARY_HEADERS) + " |",
        "| " + " | ".join("---" for _ in SUMMARY_HEADERS) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_markdown_cell(value) for value in _summary_values(row)) + " |")

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_metrics_xlsx(path: str | Path, rows: list[dict[str, Any]]) -> None:
    """Write the metrics as an Excel workbook for human review.

    Produces four sheets:
      - ``summary``: one row per benchmark run (``SUMMARY_HEADERS`` columns).
      - ``eval_metrics``: benchmark-owned quality, answer token, and timing metrics.
      - ``llm_by_task``: one row per (run, LLM task) with token and timing breakdown.
      - ``percentiles``: min/max/p50/p95 for add/search operation timings.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append(SUMMARY_HEADERS)
    for row in rows:
        summary.append([_xlsx_cell(value) for value in _summary_values(row)])

    eval_metrics_sheet = wb.create_sheet("eval_metrics")
    all_eval_rows = [[_xlsx_cell(v) for v in _eval_metrics_values(row)] for row in rows]
    non_empty_cols = [
        i for i, header in enumerate(EVAL_METRICS_HEADERS)
        if any(row[i] not in (None, "") for row in all_eval_rows)
    ]
    eval_metrics_sheet.append([EVAL_METRICS_HEADERS[i] for i in non_empty_cols])
    for row in all_eval_rows:
        eval_metrics_sheet.append([row[i] for i in non_empty_cols])

    by_task_sheet = wb.create_sheet("llm_by_task")
    by_task_headers = [
        "run_id",
        "algorithm",
        "task",
        "llm_calls",
        "llm_total_tokens",
        "llm_avg_tokens",
        "llm_total_time_ms",
    ]
    by_task_sheet.append(by_task_headers)
    for row in rows:
        clickhouse = dict(row.get("metrics", {}).get("clickhouse") or {})
        for task in clickhouse.get("llm_by_task") or []:
            if not isinstance(task, dict):
                continue
            by_task_sheet.append(
                [
                    row.get("run_id"),
                    row.get("memory_algorithm"),
                    task.get("task"),
                    _to_int(task.get("llm_call_count")),
                    _to_int(task.get("llm_total_tokens")),
                    _to_float(task.get("llm_avg_tokens")),
                    _to_float(task.get("llm_total_time_ms")),
                ]
            )
        em = dict((row.get("eval_result") or {}).get("metrics") or {})
        for stage, calls_key, tokens_key, elapsed_key in (
            ("eval.answer", "answer_llm_calls", "answer_total_tokens", "answer_elapsed_seconds"),
            ("eval.judge", "judge_llm_calls", "judge_total_tokens", None),
        ):
            calls = _to_int(em.get(calls_key))
            tokens = _to_int(em.get(tokens_key))
            elapsed_s = em.get(elapsed_key) if elapsed_key else None
            elapsed_ms = _to_float(elapsed_s * 1000) if elapsed_s is not None else None
            avg_tokens = _to_float(tokens / calls) if calls else None
            if calls:
                by_task_sheet.append(
                    [
                        row.get("run_id"),
                        row.get("memory_algorithm"),
                        stage,
                        calls,
                        tokens,
                        avg_tokens,
                        elapsed_ms,
                    ]
                )

    percentiles_sheet = wb.create_sheet("percentiles")
    percentiles_headers = [
        "run_id",
        "algorithm",
        "operation",
        "time_min_ms",
        "time_max_ms",
        "time_p50_ms",
        "time_p95_ms",
        "token_min",
        "token_max",
        "token_p50",
        "token_p95",
    ]
    percentiles_sheet.append(percentiles_headers)
    for row in rows:
        qdrant = dict(row.get("metrics", {}).get("qdrant") or {})
        clickhouse = dict(row.get("metrics", {}).get("clickhouse") or {})
        token_perc = clickhouse.get("llm_token_percentiles_by_task") or {}

        for op_name in ("add", "search"):
            time_percentiles = qdrant.get(f"{op_name}_percentiles_ms") or {}

            if op_name == "search":
                sp = clickhouse.get("search_token_percentiles") or {}
                token_min = _to_int(sp["token_min"]) if sp.get("token_min") is not None else None
                token_max = _to_int(sp["token_max"]) if sp.get("token_max") is not None else None
                token_p50 = _to_int(sp["token_p50"]) if sp.get("token_p50") is not None else None
                token_p95 = _to_int(sp["token_p95"]) if sp.get("token_p95") is not None else None
            else:
                token_stats = _aggregate_token_percentiles(token_perc, op_name)
                token_min = token_stats.get("min")
                token_max = token_stats.get("max")
                token_p50 = token_stats.get("p50")
                token_p95 = token_stats.get("p95")

            percentiles_sheet.append(
                [
                    row.get("run_id"),
                    row.get("memory_algorithm"),
                    op_name,
                    time_percentiles.get("min"),
                    time_percentiles.get("max"),
                    time_percentiles.get("p50"),
                    time_percentiles.get("p95"),
                    token_min,
                    token_max,
                    token_p50,
                    token_p95,
                ]
            )
        qa_series = _eval_qa_token_series(row)
        em = dict((row.get("eval_result") or {}).get("metrics") or {})
        for stage, elapsed_key in (
            ("answer", "answer_elapsed_seconds"),
            ("judge", None),
        ):
            token_vals = qa_series.get(stage) or []
            if not token_vals:
                continue
            tok_stats = _percentile_stats(token_vals)
            calls = len(token_vals)
            elapsed_s = em.get(elapsed_key) if elapsed_key else None
            avg_time_ms = _to_float(elapsed_s * 1000 / calls) if (elapsed_s and calls) else None
            percentiles_sheet.append(
                [
                    row.get("run_id"),
                    row.get("memory_algorithm"),
                    f"eval.{stage}",
                    None,
                    None,
                    avg_time_ms,
                    None,
                    tok_stats["min"],
                    tok_stats["max"],
                    tok_stats["p50"],
                    tok_stats["p95"],
                ]
            )

    for sheet in (summary, eval_metrics_sheet, by_task_sheet, percentiles_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)


def _xlsx_cell(value: Any) -> Any:
    """Coerce a metrics value to an Excel-friendly scalar (keep numbers numeric)."""
    if value is None or isinstance(value, (int, float, str)):
        return value
    return str(value)


def _markdown_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ")


def config_from_args(args: argparse.Namespace) -> MetricsCollectorConfig:
    """Build collector config from CLI args and environment defaults."""
    return MetricsCollectorConfig(
        qdrant=QdrantCollectorConfig(
            url=args.qdrant_url or os.getenv("MINDMEMOS_QDRANT_URL") or DEFAULT_QDRANT_URL,
            api_key=args.qdrant_api_key or os.getenv("MINDMEMOS_QDRANT_API_KEY"),
            memory_collection=args.memory_collection
            or os.getenv("MINDMEMOS_QDRANT_MEMORY_COLLECTION")
            or DEFAULT_MEMORY_COLLECTION,
            add_record_collection=args.add_record_collection
            or os.getenv("MINDMEMOS_QDRANT_ADD_RECORD_COLLECTION")
            or DEFAULT_ADD_RECORD_COLLECTION,
            search_record_collection=args.search_record_collection
            or os.getenv("MINDMEMOS_QDRANT_SEARCH_RECORD_COLLECTION")
            or DEFAULT_SEARCH_RECORD_COLLECTION,
            page_size=args.page_size,
        ),
        clickhouse=ClickHouseCollectorConfig(
            url=args.clickhouse_url or os.getenv("MINDMEMOS_CLICKHOUSE_HTTP_URL") or DEFAULT_CLICKHOUSE_URL,
            database=args.clickhouse_database or os.getenv("MINDMEMOS_CLICKHOUSE_DB") or DEFAULT_CLICKHOUSE_DATABASE,
            table=args.clickhouse_table or os.getenv("MINDMEMOS_CLICKHOUSE_TRACES_TABLE") or DEFAULT_CLICKHOUSE_TABLE,
            user=args.clickhouse_user or os.getenv("MINDMEMOS_CLICKHOUSE_USER") or DEFAULT_CLICKHOUSE_USER,
            password=args.clickhouse_password
            or os.getenv("MINDMEMOS_CLICKHOUSE_PASSWORD")
            or DEFAULT_CLICKHOUSE_PASSWORD,
            timeout_seconds=args.clickhouse_timeout,
            enabled=not args.skip_clickhouse,
            strict=args.strict_clickhouse,
        ),
    )


def build_arg_parser() -> argparse.ArgumentParser:
    """Build the metrics collector CLI parser."""
    parser = argparse.ArgumentParser(description="Collect benchmark matrix Qdrant and ClickHouse metrics.")
    parser.add_argument("--manifest", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--qdrant-url")
    parser.add_argument("--qdrant-api-key")
    parser.add_argument("--memory-collection")
    parser.add_argument("--add-record-collection")
    parser.add_argument("--search-record-collection")
    parser.add_argument("--page-size", type=int, default=256)
    parser.add_argument("--clickhouse-url")
    parser.add_argument("--clickhouse-database")
    parser.add_argument("--clickhouse-table")
    parser.add_argument("--clickhouse-user")
    parser.add_argument("--clickhouse-password")
    parser.add_argument("--clickhouse-timeout", type=float, default=10.0)
    parser.add_argument("--skip-clickhouse", action="store_true")
    parser.add_argument("--strict-clickhouse", action="store_true")
    parser.add_argument("--table-output", help="Optional Markdown table output for compact human review.")
    parser.add_argument("--xlsx-output", help="Optional Excel (.xlsx) output for human review.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    args = build_arg_parser().parse_args(argv)
    asyncio.run(
        collect_benchmark_metrics(
            args.manifest,
            args.output,
            config_from_args(args),
            args.table_output,
            args.xlsx_output,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
